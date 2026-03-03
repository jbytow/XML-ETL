"""
XML to Excel Export.

Scans input/xml/ for all .xml files (Comarch Optima invoices),
extracts header + line item fields, and writes to Excel.
Each line item (POZYCJA) becomes one row, with header data repeated.
"""

import sys
import logging
from datetime import datetime
from pathlib import Path

from lxml import etree
from openpyxl import Workbook, load_workbook
from tqdm import tqdm

import config

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s: %(message)s",
)
log = logging.getLogger(__name__)

NS = config.XML_NAMESPACE


def extract_field(element, xpath: str) -> str | None:
    """Extract text from the first element matching the XPath."""
    results = element.xpath(xpath, namespaces=NS)
    if not results:
        return None
    el = results[0]
    return el.text if hasattr(el, "text") else str(el)


def parse_xml_file(filepath: Path) -> list[dict]:
    """Parse an invoice XML and return one dict per line item."""
    try:
        tree = etree.parse(str(filepath))
    except etree.XMLSyntaxError as e:
        log.warning("Skipping %s - XML parse error: %s", filepath.name, e)
        return []

    root = tree.getroot()

    # Find DOKUMENT element
    docs = root.xpath("ns:DOKUMENT", namespaces=NS)
    if not docs:
        # Maybe root IS the document
        docs = [root]
    doc = docs[0]

    # Extract header fields (once per file)
    header = {"filename": filepath.name}
    for field in config.HEADER_FIELDS:
        header[field["column"]] = extract_field(doc, field["xpath"])

    # Extract line items
    items = doc.xpath(config.ITEM_XPATH, namespaces=NS)
    if not items:
        log.warning("No line items found in %s", filepath.name)
        return [header]  # still export the header row

    rows = []
    for item in items:
        row = dict(header)  # copy header fields
        for field in config.ITEM_FIELDS:
            row[field["column"]] = extract_field(item, field["xpath"])
        rows.append(row)

    return rows


def main():
    xml_dir = config.XML_DIR
    output_dir = config.OUTPUT_DIR

    if not xml_dir.exists():
        log.error("XML directory not found: %s", xml_dir)
        sys.exit(1)

    xml_files = sorted(xml_dir.glob("*.xml"))
    if not xml_files:
        log.error("No .xml files found in %s", xml_dir)
        sys.exit(1)

    log.info("Found %d XML files in %s", len(xml_files), xml_dir)

    # Parse all files
    all_rows = []
    errors = 0
    for filepath in tqdm(xml_files, desc="Parsing XML files"):
        rows = parse_xml_file(filepath)
        if rows:
            all_rows.extend(rows)
        else:
            errors += 1

    if not all_rows:
        log.error("No data was parsed successfully.")
        sys.exit(1)

    # Load kartoteka mapping (Kod -> Kod Zaktualizowany)
    code_mapping = {}
    kartoteka = config.KARTOTEKA_PATH
    if kartoteka.exists():
        log.info("Loading code mapping from %s", kartoteka.name)
        kwb = load_workbook(str(kartoteka), read_only=True, data_only=True)
        kws = kwb.active
        kheaders = [c.value for c in next(kws.iter_rows(min_row=1, max_row=1))]
        if "Kod" in kheaders and "Kod Zaktualizowany" in kheaders:
            kid = kheaders.index("Kod")
            nid = kheaders.index("Kod Zaktualizowany")
            for krow in kws.iter_rows(min_row=2, values_only=True):
                if krow[kid] and krow[nid]:
                    code_mapping[str(krow[kid]).strip()] = str(krow[nid]).strip()
            log.info("Loaded %d code mappings", len(code_mapping))
        else:
            log.warning("Kartoteka missing 'Kod' or 'Kod Zaktualizowany' column, skipping mapping")
        kwb.close()
    else:
        log.info("No kartoteka file found at %s, skipping mapping", kartoteka)

    # Add mapped code to each row
    for row in all_rows:
        kod = row.get("TowarKod")
        row["KodZaktualizowany"] = code_mapping.get(kod) if kod else None

    # Build column list
    header_cols = [f["column"] for f in config.HEADER_FIELDS]
    item_cols = [f["column"] for f in config.ITEM_FIELDS]
    columns = ["filename"] + header_cols + item_cols + ["KodZaktualizowany"]

    # Write Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Faktury"
    ws.append(columns)

    for row in all_rows:
        ws.append([row.get(col) for col in columns])

    # Auto-size columns (approximate)
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for ws_row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in ws_row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 2, 50)

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"xml_export_{timestamp}.xlsx"
    wb.save(str(output_file))

    log.info(
        "Exported %d rows (%d files) to %s",
        len(all_rows), len(xml_files) - errors, output_file,
    )
    if errors:
        log.warning("Skipped %d files due to errors", errors)


if __name__ == "__main__":
    main()
