"""
Excel-driven XML Update.

Reads a product mapping Excel file with "Kod" (current) and "Kod Zaktualizowany" (new)
columns, then scans all XML invoices and replaces matching product codes.
Modified files are saved to output/xml/ (originals untouched).

Usage:
    python excel_to_xml.py path/to/kartoteka.xlsx
"""

import sys
import logging
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook
from tqdm import tqdm

import config

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s: %(message)s",
)
log = logging.getLogger(__name__)

NS = config.XML_NAMESPACE

# --- Excel column names for the mapping file ---
EXCEL_KEY_COLUMN = "Kod"                  # current product code (obecny kod)
EXCEL_NEW_COLUMN = "Kod Zaktualizowany"   # new product code (nowy kod)

# --- XPath to product code elements to update (relative to POZYCJA) ---
CODE_XPATHS = [
    "ns:TOWAR/ns:KOD",
    "ns:TOWAR/ns:NUMER_KATALOGOWY",
]


def load_code_mapping(excel_path: Path) -> dict[str, str]:
    """Load Kod -> Kod Zaktualizowany mapping from Excel. Skip rows without update."""
    wb = load_workbook(str(excel_path), read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    if EXCEL_KEY_COLUMN not in headers:
        log.error("Column '%s' not found. Available: %s", EXCEL_KEY_COLUMN, headers)
        sys.exit(1)
    if EXCEL_NEW_COLUMN not in headers:
        log.error("Column '%s' not found. Available: %s", EXCEL_NEW_COLUMN, headers)
        sys.exit(1)

    key_idx = headers.index(EXCEL_KEY_COLUMN)
    new_idx = headers.index(EXCEL_NEW_COLUMN)

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[key_idx] or not row[new_idx]:
            continue
        old_code = str(row[key_idx]).strip()
        new_code = str(row[new_idx]).strip()
        if old_code and new_code and old_code != new_code:
            mapping[old_code] = new_code

    wb.close()
    return mapping


def update_xml_file(xml_path: Path, output_path: Path, code_mapping: dict[str, str]) -> list[str]:
    """
    Scan all POZYCJA elements in the XML file and replace product codes
    that exist in the mapping. Saves the result to output_path.
    Returns list of change descriptions.
    """
    tree = etree.parse(str(xml_path))
    root = tree.getroot()

    docs = root.xpath("ns:DOKUMENT", namespaces=NS)
    doc = docs[0] if docs else root

    items = doc.xpath(config.ITEM_XPATH, namespaces=NS)
    changes = []

    for item_el in items:
        # Read current KOD to check if it's in the mapping
        kod_results = item_el.xpath("ns:TOWAR/ns:KOD", namespaces=NS)
        if not kod_results:
            continue

        old_code = kod_results[0].text
        if old_code not in code_mapping:
            continue

        new_code = code_mapping[old_code]

        # Update all code-related XPaths
        for xpath in CODE_XPATHS:
            results = item_el.xpath(xpath, namespaces=NS)
            if not results:
                continue
            el = results[0]
            if el.text == old_code:
                el.text = new_code
                changes.append(f"{xpath}: '{old_code}' -> '{new_code}'")

    # Always write to output (copy even if no changes, so output has complete set)
    tree.write(
        str(output_path),
        xml_declaration=True,
        encoding=tree.docinfo.encoding or "UTF-8",
        pretty_print=True,
    )

    return changes


def main():
    if len(sys.argv) < 2:
        print("Usage: python excel_to_xml.py <path_to_excel_file>")
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    if not excel_path.exists():
        log.error("Excel file not found: %s", excel_path)
        sys.exit(1)

    xml_dir = config.XML_DIR
    if not xml_dir.exists():
        log.error("XML directory not found: %s", xml_dir)
        sys.exit(1)

    # Load code mapping from Excel
    code_mapping = load_code_mapping(excel_path)
    log.info("Loaded %d code mappings from %s", len(code_mapping), excel_path.name)

    if not code_mapping:
        log.info("No codes to update (all 'Kod Zaktualizowany' values are empty or identical to 'Kod').")
        sys.exit(0)

    # Show a few examples
    examples = list(code_mapping.items())[:5]
    for old, new in examples:
        log.info("  %s -> %s", old, new)
    if len(code_mapping) > 5:
        log.info("  ... and %d more", len(code_mapping) - 5)

    # Process all XML files
    xml_files = sorted(xml_dir.glob("*.xml"))
    if not xml_files:
        log.error("No .xml files found in %s", xml_dir)
        sys.exit(1)

    log.info("Scanning %d XML files...", len(xml_files))

    output_xml_dir = config.OUTPUT_DIR / "xml"
    output_xml_dir.mkdir(parents=True, exist_ok=True)

    total_changes = 0
    files_updated = 0

    for xml_path in tqdm(xml_files, desc="Updating XML files"):
        output_path = output_xml_dir / xml_path.name
        try:
            changes = update_xml_file(xml_path, output_path, code_mapping)
            if changes:
                files_updated += 1
                total_changes += len(changes)
                for c in changes:
                    log.info("  %s: %s", xml_path.name, c)
        except Exception as e:
            log.error("Error processing %s: %s", xml_path.name, e)

    log.info("Done. %d files updated, %d code replacements total.", files_updated, total_changes)
    log.info("Output XML files saved to %s", output_xml_dir)


if __name__ == "__main__":
    main()
